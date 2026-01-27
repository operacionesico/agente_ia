"""
Lector de Información de la Empresa
Lee automáticamente documentos PDF, Word, Excel de la carpeta de información
y extrae el texto para usarlo como contexto en las generaciones de IA.
"""

import os
import glob
from PyPDF2 import PdfReader
from docx import Document
import openpyxl

def extraer_texto_pdf(ruta_archivo):
    """Extrae texto de un archivo PDF"""
    try:
        reader = PdfReader(ruta_archivo)
        texto = ""
        for pagina in reader.pages:
            texto += pagina.extract_text() + "\n"
        return texto.strip()
    except Exception as e:
        return f"[Error leyendo PDF: {e}]"

def extraer_texto_word(ruta_archivo):
    """Extrae texto de un archivo Word (.docx)"""
    try:
        doc = Document(ruta_archivo)
        texto = ""
        for parrafo in doc.paragraphs:
            texto += parrafo.text + "\n"
        return texto.strip()
    except Exception as e:
        return f"[Error leyendo Word: {e}]"

def extraer_texto_excel(ruta_archivo):
    """Extrae texto de un archivo Excel (.xlsx)"""
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        texto = ""
        
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            texto += f"\n--- Hoja: {hoja} ---\n"
            
            for fila in ws.iter_rows(values_only=True):
                valores = [str(v) for v in fila if v is not None]
                if valores:
                    texto += " | ".join(valores) + "\n"
        
        wb.close()
        return texto.strip()
    except Exception as e:
        return f"[Error leyendo Excel: {e}]"

def extraer_texto_txt(ruta_archivo):
    """Extrae texto de un archivo de texto plano"""
    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception as e:
        return f"[Error leyendo TXT: {e}]"

def leer_informacion_empresa(carpeta="4. INFORMACION EMPRESA"):
    """
    Lee TODOS los archivos de la carpeta de información de la empresa
    y extrae su contenido como texto.
    
    Args:
        carpeta: Ruta a la carpeta con información
    
    Returns:
        String con todo el contenido extraído, formateado
    """
    if not os.path.exists(carpeta):
        return None
    
    archivos_pdf = glob.glob(os.path.join(carpeta, "*.pdf"))
    archivos_word = glob.glob(os.path.join(carpeta, "*.docx"))
    archivos_excel = glob.glob(os.path.join(carpeta, "*.xlsx"))
    archivos_txt = glob.glob(os.path.join(carpeta, "*.txt"))
    
    total_archivos = len(archivos_pdf) + len(archivos_word) + len(archivos_excel) + len(archivos_txt)
    
    if total_archivos == 0:
        return None
    
    print(f"📂 Leyendo {total_archivos} documentos de información empresarial...")
    
    contenido_total = "\n" + "="*80 + "\n"
    contenido_total += "INFORMACIÓN ADICIONAL DE LA EMPRESA (usar para generar contenido realista)\n"
    contenido_total += "="*80 + "\n\n"
    
    for archivo in archivos_pdf:
        nombre = os.path.basename(archivo)
        print(f"   📄 {nombre}")
        texto = extraer_texto_pdf(archivo)
        contenido_total += f"\n--- Archivo: {nombre} ---\n{texto}\n\n"
    
    for archivo in archivos_word:
        nombre = os.path.basename(archivo)
        print(f"   📝 {nombre}")
        texto = extraer_texto_word(archivo)
        contenido_total += f"\n--- Archivo: {nombre} ---\n{texto}\n\n"
    
    for archivo in archivos_excel:
        nombre = os.path.basename(archivo)
        print(f"   📊 {nombre}")
        texto = extraer_texto_excel(archivo)
        contenido_total += f"\n--- Archivo: {nombre} ---\n{texto}\n\n"
    
    for archivo in archivos_txt:
        nombre = os.path.basename(archivo)
        print(f"   📃 {nombre}")
        texto = extraer_texto_txt(archivo)
        contenido_total += f"\n--- Archivo: {nombre} ---\n{texto}\n\n"
    
    contenido_total += "="*80 + "\n"
    
    print(f"✅ Información empresarial cargada ({total_archivos} archivos)")
    
    return contenido_total

if __name__ == "__main__":
    info = leer_informacion_empresa()
    if info:
        print(info[:500])
    else:
        print("No se encontró información empresarial")
