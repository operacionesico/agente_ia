"""
Script para inyectar datos y contenido IA en plantillas de Excel.
Lee configuraciones desde Excel en '1. Data/', procesa plantillas en '2. Plantilla/',
y guarda resultados en '3. Inyectado/'.
"""

import os
import glob
import openpyxl
import re
from datetime import datetime
from gemini_client import GeminiClient
from config_auditoria import generar_contexto_base
from lector_informacion import leer_informacion_empresa

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, '1. Data')
PLANTILLA_DIR = os.path.join(BASE_DIR, '2. Plantilla')
INYECTADO_DIR = os.path.join(BASE_DIR, '3. Inyectado')

def leer_datos_excel():
    """
    Busca el primer archivo Excel en '1. Data/' y lee las configuraciones.
    
    Estructura (3 columnas):
    - Columna A: Campo (nombre de la variable)
    - Columna B: Valor (dato estático O prompt si es IA)
    - Columna C: Campo Generado (etiqueta)
    
    Detección automática:
    - Si Campo Generado empieza con {{IA: → Valor contiene el prompt
    - Si no → Valor contiene el dato estático
    
    Retorna dos diccionarios: datos_estaticos y datos_ia
    """
    archivos_excel = glob.glob(os.path.join(DATA_DIR, '*.xlsx'))
    
    if not archivos_excel:
        print(f"❌ No se encontró ningún archivo Excel en la carpeta '{DATA_DIR}'")
        return None, None, None
    
    archivo_excel = archivos_excel[0]
    print(f"📄 Leyendo datos desde: {os.path.basename(archivo_excel)}")
    
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    ws = wb.active
    
    datos_estaticos = {}
    datos_ia = {}
    
    print("   Leyendo filas...")
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]: continue
        
        campo = str(row[0]).strip()
        valor = row[1] if row[1] is not None else ""
        campo_generado = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        
        if campo_generado.startswith("{{IA:"):
            datos_ia[campo] = valor
            print(f"   - Fila {i}: {campo} [IA] → Prompt configurado")
        else:
            datos_estaticos[campo] = valor
            print(f"   - Fila {i}: {campo} [STATIC] → {str(valor)[:50]}...")
    
    normas = []
    for campo, valor in datos_estaticos.items():
        if 'NORMA' in campo.upper() and valor and not campo.endswith('_RESUMEN'):
            normas.append(str(valor))
    
    wb.close()
    
    print(f"\n✅ Datos cargados: {len(datos_estaticos)} estáticos, {len(datos_ia)} IA")
    if normas:
        print(f"📋 Normas detectadas: {', '.join(normas)}")
    
    return datos_estaticos, datos_ia, normas

def procesar_celda(celda, datos_estaticos, datos_ia, cliente_gemini, contexto_sistema, memoria_respuestas):
    """
    Procesa una celda individual de Excel, reemplazando etiquetas.
    
    Args:
        celda: Objeto Cell de openpyxl
        datos_estaticos: Diccionario con datos estáticos
        datos_ia: Diccionario con prompts de IA
        cliente_gemini: Cliente de Gemini API
        contexto_sistema: Contexto base del sistema
        memoria_respuestas: Lista con respuestas previas de IA
    """
    if celda.value is None or not isinstance(celda.value, str):
        return
    
    texto = str(celda.value)
    cambios = False
    
    for campo, valor in datos_estaticos.items():
        etiqueta = f"{{{{{campo}}}}}"
        if etiqueta in texto:
            texto = texto.replace(etiqueta, str(valor))
            cambios = True
    
    etiquetas_ia = re.findall(r'\{\{IA:([^}]+)\}\}', texto)
    
    for nombre_prompt in etiquetas_ia:
        if nombre_prompt in datos_ia:
            prompt = datos_ia[nombre_prompt]
            
            print(f"   🤖 Generando contenido IA para: {nombre_prompt}")
            
            respuesta = cliente_gemini.generar_texto(prompt, datos_estaticos, contexto_sistema)
            
            memoria_respuestas.append(f"[{nombre_prompt}]\n{respuesta}")
            
            texto = texto.replace(f"{{{{IA:{nombre_prompt}}}}}", respuesta)
            cambios = True
            
            print(f"      ✅ Generado ({len(respuesta)} caracteres)")
    
    if cambios:
        texto = re.sub(r',\s*,', ',', texto)
        texto = re.sub(r'\s+,', ',', texto)
        texto = re.sub(r',\s*\.', '.', texto)
        texto = re.sub(r'\s{2,}', ' ', texto)
        
        celda.value = texto.strip()

def procesar_excel():
    """
    Función principal que procesa todas las plantillas Excel.
    """
    print("\n" + "="*80)
    print("🚀 INICIANDO PROCESAMIENTO DE PLANTILLAS EXCEL")
    print("="*80 + "\n")
    
    os.makedirs(INYECTADO_DIR, exist_ok=True)
    
    datos_estaticos, datos_ia, normas = leer_datos_excel()
    
    if datos_estaticos is None:
        print("❌ No se pudieron cargar datos. Verifica la carpeta '1. Data'")
        return
    
    contexto_empresa = leer_informacion_empresa()
    
    contexto_sistema = generar_contexto_base(normas, datos_estaticos, None, contexto_empresa)
    
    ruta_contexto = os.path.join(INYECTADO_DIR, 'CONTEXTO_IA_EXCEL.txt')
    try:
        with open(ruta_contexto, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("CONTEXTO COMPLETO ENVIADO A LA IA (EXCEL)\n")
            f.write(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write("="*80 + "\n\n")
            f.write(contexto_sistema)
        print(f"💾 Contexto guardado en: {os.path.basename(ruta_contexto)}")
    except Exception as e:
        print(f"⚠️  No se pudo guardar contexto: {e}")
    
    if normas:
        print(f"📋 Normas detectadas para auditoría: {', '.join(normas)}")
    else:
        print("⚠️  No se detectaron normas en el Excel (campos NORMA1, NORMA2, etc.)")
    
    print(f"📊 Datos cargados: {len(datos_estaticos)} variables estáticas, {len(datos_ia)} prompts de IA")
    
    todas_plantillas = glob.glob(os.path.join(PLANTILLA_DIR, '*.xlsx'))
    plantillas = [p for p in todas_plantillas if not os.path.basename(p).startswith('~$')]
    
    if not plantillas:
        print(f"❌ No se encontraron plantillas .xlsx en '{PLANTILLA_DIR}'")
        return
    
    print(f"\n📁 Plantillas encontradas: {len(plantillas)}")
    for p in plantillas:
        print(f"   - {os.path.basename(p)}")
    
    print("\n🤖 Inicializando cliente Gemini...")
    try:
        cliente_gemini = GeminiClient()
        print("   ✅ Cliente Gemini listo")
    except Exception as e:
        print(f"   ❌ Error al inicializar Gemini: {e}")
        return
    
    print("\n" + "="*80)
    print("📝 PROCESANDO PLANTILLAS")
    print("="*80 + "\n")
    
    for idx, plantilla_path in enumerate(plantillas, 1):
        nombre_plantilla = os.path.basename(plantilla_path)
        print(f"\n[{idx}/{len(plantillas)}] Procesando: {nombre_plantilla}")
        print("-" * 60)
        
        try:
            wb = openpyxl.load_workbook(plantilla_path)
            
            memoria_respuestas = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n   📄 Hoja: {sheet_name}")
                
                for row in ws.iter_rows():
                    for celda in row:
                        procesar_celda(celda, datos_estaticos, datos_ia, cliente_gemini, contexto_sistema, memoria_respuestas)
            
            nombre_salida = nombre_plantilla.lstrip('_')
            ruta_salida = os.path.join(INYECTADO_DIR, nombre_salida)
            
            wb.save(ruta_salida)
            print(f"\n   ✅ Guardado: {nombre_salida}")
            
            if memoria_respuestas:
                nombre_memoria = f"MEMORIA_{nombre_salida.replace('.xlsx', '.txt')}"
                ruta_memoria = os.path.join(INYECTADO_DIR, nombre_memoria)
                with open(ruta_memoria, 'w', encoding='utf-8') as f:
                    f.write("="*80 + "\n")
                    f.write(f"MEMORIA DE RESPUESTAS IA - {nombre_plantilla}\n")
                    f.write(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                    f.write("="*80 + "\n\n")
                    f.write("\n\n".join(memoria_respuestas))
                print(f"   💾 Memoria guardada: {nombre_memoria}")
            
        except Exception as e:
            print(f"   ❌ Error procesando {nombre_plantilla}: {e}")
            continue
    
    print("\n" + "="*80)
    print("✅ PROCESAMIENTO COMPLETADO")
    print("="*80)
    print(f"\n📂 Documentos generados en: {INYECTADO_DIR}")

if __name__ == "__main__":
    procesar_excel()
